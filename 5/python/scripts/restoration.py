import numpy as np 
import openpyxl

def restoration(path_to_xlsx, workbook, sheet_name, W, NM_x, orig_vactor):
    
    sheet = workbook.create_sheet(title=sheet_name)

    err_cnt = 0

    for cnt in range(225):

        noise_vector = NM_x[cnt]
        row_offset = cnt * 9

        for row in range(7):
            for column in range(4):
                sheet.cell(row=row+1+row_offset,column=column+1,value=noise_vector[row*4+column]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                if noise_vector[row*4+column] == 1:
                    sheet.cell(row=row+1+row_offset,column=column+1).fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        restor_vector = np.dot(W,noise_vector)
        bin_restor_vector = np.where(restor_vector>0,1,-1)
        if not np.all(bin_restor_vector == orig_vactor):
            err_cnt += 1

        for row in range(7):
            for column in range(4):
                sheet.cell(row=row+1+row_offset,column=column+6,value=restor_vector[row*4+column]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                if restor_vector[row*4+column] >= 0:
                    sheet.cell(row=row+1+row_offset,column=column+6).fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    sheet.cell(row=1,column=11,value=f'Кількість помилок - {(err_cnt*100/225):.0f} %').alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet.cell(row=1,column=11).fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    sheet.column_dimensions[openpyxl.utils.get_column_letter(11)].width = 25

    workbook.save(path_to_xlsx)
    return restor_vector

