import numpy as np 
import openpyxl


def print_matrix_to_excle_with_color(sheet, offset_row, offset_column, M):
     for row in range(M.shape[0]):
        for column in range(M.shape[1]):
            sheet.cell(row=row+1+offset_row,column=column+1+offset_column,value=M[row,column]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            if M[row,column] >= 0:
                sheet.cell(row=row+1+offset_row,column=column+1+offset_column,).fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')



def vector_to_matrix(vector):
    matrix = np.zeros((7,4))
    for cnt in range(7):
        matrix[cnt] = vector[cnt*4:(cnt+1)*4]
    return matrix



def hopfield_recall(path_to_xlsx, workbook, sheet_name, W, NM_x, orig_vactor, iter_max):

    sheet = workbook.create_sheet(title=sheet_name)
    err_cnt = 0    

    # Print noise vector
    for cnt in range(225):
        print_matrix_to_excle_with_color(sheet, cnt*10, 0, vector_to_matrix(NM_x[cnt]))


    for iter_cnt in range(iter_max):

        for cnt in range(225):
            restor_vector = (np.dot(W, restor_vector))
            bin_restor_vector = np.where(restor_vector>0,1,-1)
            if not np.all(bin_restor_vector == orig_vactor):
                err_cnt += 1

    # for column_cnt in range(iter_max):

    #     column_offset = column_cnt * 6

    #     for row_cnt in range(225):

    #         # Set param for some noise vector
    #         noise_vector = NM_x[row_cnt]
    #         restor_vector = noise_vector
    #         row_offset = row_cnt * 9

             
            
            
            
    #         restor_vector = (np.dot(W, restor_vector))
    #         bin_restor_vector = np.where(restor_vector>0,1,-1)
    #         if not np.all(bin_restor_vector == orig_vactor):
    #             err_cnt += 1

    #         for row in range(7):
    #             for column in range(4):
    #                 sheet.cell(row=row+1+row_offset,column=column+6,value=restor_vector[row*4+column]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    #                 if restor_vector[row*4+column] >= 0:
    #                     sheet.cell(row=row+1+row_offset,column=column+6).fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


        
    workbook.save(path_to_xlsx)