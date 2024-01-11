import numpy as np 
import openpyxl


def get_XTX(path_to_xlsx, workbook, sheet_name, symbol_vector):
    
    sheet = workbook.create_sheet(title=sheet_name)

    for cnt in range(28):
        sheet.cell(row=1, column=cnt+2, value=symbol_vector[cnt]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet.cell(row=cnt+2, column=1, value=symbol_vector[cnt]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet.cell(row=1, column=cnt+2).fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        sheet.cell(row=cnt+2, column=1).fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    XTX = np.zeros((28,28))

    for y in range(28):
        for x in range(28):
            XTX[y,x] = symbol_vector[y] * symbol_vector[x]
            sheet.cell(row=y+2, column=x+2, value=XTX[y,x]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    workbook.save(path_to_xlsx)

    return XTX
