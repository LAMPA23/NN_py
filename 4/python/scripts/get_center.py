import numpy as np 
import openpyxl



# Get center of one symbol
def get_center(noise_matrix_one_symbol):
    center = np.zeros(28)
    for index in range(28):
        for row in range(225):
            center[index] += noise_matrix_one_symbol[row, index]
        center[index] /= 225
    return center



# Get centers of all symbols lile tuple
def get_centers(path_to_xlsx, workbook, noise_matrix_all):
    
    # Get enters of symbols
    center_of_symbol_1 = get_center(noise_matrix_all[0:225,:])
    center_of_symbol_2 = get_center(noise_matrix_all[225:450,:])
    center_of_symbol_3 = get_center(noise_matrix_all[450:675,:])


    # Create new sheet
    sheet = workbook.create_sheet(title='Центроїди')


    # Centring a text in cell
    for row in [0,2,4]:
        for column in range(29):
            sheet.cell(row=row+1, column=column+1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')


    # Add color to name of symbols
    sheet.cell(row=1, column=1, value='Символ №1').fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    sheet.cell(row=3, column=1, value='Символ №2').fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    sheet.cell(row=5, column=1, value='Символ №3').fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")


    # Add color and value to symbols center table 
    for column in range(28):
        sheet.cell(row=1, column=column+2, value=center_of_symbol_1[column]).fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        sheet.cell(row=3, column=column+2, value=center_of_symbol_2[column]).fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        sheet.cell(row=5, column=column+2, value=center_of_symbol_3[column]).fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")


    # Save data to Excel
    workbook.save(path_to_xlsx)


    # Return three centers
    return center_of_symbol_1, center_of_symbol_2, center_of_symbol_3
