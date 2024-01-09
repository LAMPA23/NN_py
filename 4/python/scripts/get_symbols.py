import numpy as np 
import openpyxl

def get_symbols(path_to_xlsx):

    # Create work_book and sheet
    work_book = openpyxl.load_workbook(path_to_xlsx)
    sheet = work_book.active


    # Get vector of symbols
    symbol_1 = np.zeros(28)
    symbol_2 = np.zeros(28)
    symbol_3 = np.zeros(28)

    cnt_arr = 0
    for row_cnt in range(1,8):
        for column_cnt in range(1,5):
            symbol_1[cnt_arr] = sheet.cell(row=row_cnt, column=column_cnt).value
            cnt_arr += 1

    cnt_arr = 0
    for row_cnt in range(11,18):
        for column_cnt in range(1,5):
            symbol_2[cnt_arr] = sheet.cell(row=row_cnt, column=column_cnt).value
            cnt_arr += 1

    cnt_arr = 0
    for row_cnt in range(21,28):
        for column_cnt in range(1,5):
            symbol_3[cnt_arr] = sheet.cell(row=row_cnt, column=column_cnt).value
            cnt_arr += 1

    return symbol_1, symbol_2, symbol_3