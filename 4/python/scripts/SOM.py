from minisom import MiniSom
import openpyxl
import numpy as np

def create_SOM(path_to_xlsx, workbook, sheet_name, mixed_noise_matrix):
    som = MiniSom(3, 1, input_len=mixed_noise_matrix.shape[1], sigma=1.0, learning_rate=0.5)
    som.train_random(mixed_noise_matrix, 1000) 
    som_map = som.get_weights()

    
    # Create new sheet
    sheet = workbook.create_sheet(title=sheet_name)

    # Centring a text in cell
    for row in range(8):
        for column in range(29):
            sheet.cell(row=row+1, column=column+1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')


    # Add color to name of symbols
    sheet.cell(row=1, column=1, value='Центроїди отримані за допомогою НМ КК').fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    sheet.cell(row=2, column=1, value='Символ №1').fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    sheet.cell(row=3, column=1, value='Символ №2').fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    sheet.cell(row=4, column=1, value='Символ №3').fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")


    # Add color and value to symbols center table 
    for column in range(28):
        sheet.cell(row=2, column=column+2, value=som_map[0,0][column]).fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        sheet.cell(row=3, column=column+2, value=som_map[1,0][column]).fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        sheet.cell(row=4, column=column+2, value=som_map[2,0][column]).fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")



    # Save data to Excel   
    workbook.save(path_to_xlsx)

    return som_map[0,0], som_map[1,0], som_map[2,0]
