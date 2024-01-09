import numpy as np 
import openpyxl



def create_mixed_data(path_to_xlsx, workbook, sheet_name, noise_matrix):

    # Get mixed indices
    shuffled_indices = np.random.permutation(noise_matrix.shape[0])
    
    # Create target data like matrix
    target_data = np.zeros((675,3))
    target_data[0:225,0] = 1
    target_data[225:450,1] = 1
    target_data[450:675,2] = 1

    # Create mixed noise matrix and traget data
    mixed_matrix = noise_matrix[shuffled_indices,:]
    mixed_target_data = target_data[shuffled_indices,:]

    # Excel preparation
    sheet = workbook.create_sheet(title=sheet_name)

    # Centring text in cell
    for row in range(675):
        for column in range(35):
            sheet.cell(row=row+1, column=column+1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # Excel print number of row (in not mixed data)
    for cnt in range(675):
        sheet.cell(row=cnt+1, column=1, value=shuffled_indices[cnt]).fill = openpyxl.styles.PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")
    
    # Excel print mixed noise matrix
    for row in range(675):
        for column in range(28):
            if noise_matrix[row, column] == 1:
                sheet.cell(row=row+1, column=column+2, value=1).fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            else:
                sheet.cell(row=row+1, column=column+2, value=-1)

    # Excel print mixed traget data
    for row in range(675):
        for column in range(3):
            sheet.cell(row=row+1, column=column+31, value=mixed_target_data[row,column]).fill = openpyxl.styles.PatternFill(start_color="009900", end_color="009900", fill_type="solid")
            

    # Save data to Excel
    workbook.save(path_to_xlsx)

    return mixed_matrix, mixed_target_data
