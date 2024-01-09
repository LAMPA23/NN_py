import numpy as np 
import openpyxl


def create_noise_matrix(path_to_xlsx, workbook, symbols):

    # Excel preparation
    sheet = workbook.create_sheet(title='Зашумлена матриця')
    

    # Creatin training matrix 675x28
    M = np.zeros((675,28)).astype(int)
    M[0:225,:] = symbols[0]
    M[225:450,:] = symbols[1]
    M[450:675,:] = symbols[2]


    # Add clear matrix to xlsx
    for row in range(M.shape[0]):
        for column in range(M.shape[1]):
            sheet.cell(row=row+1, column=column+1, value=M[row,column]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            if M[row,column] == 1:
                sheet.cell(row=row+1, column=column+1).fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


    # Add noise to clear matrix
    for cnt_1 in range(8):
        for cnt_2 in range(28):
            inverting_vector = np.array(range(cnt_2, cnt_2 + cnt_1 + 1)).astype(int) % 28
            rows = cnt_1 * 28 + cnt_2 + 1, cnt_1 * 28 + cnt_2 + 226, cnt_1 * 28 + cnt_2 + 451
            for row in rows:
                for column in inverting_vector:
                    M[row, column] = 1 if M[row, column] == -1 else -1
                    for column in inverting_vector:
                        sheet.cell(row=row+1, column=column+1, value=M[row, column]).fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                

    # Save data to Excel
    workbook.save(path_to_xlsx)


    # Return noise matrix
    return M





def create_target_ans_for_noise_matrix(path_to_xlsx, workbook):
    
    # Create target data like matrix
    target_data = np.zeros((675,28))
    target_data[0:225,0] = 1
    target_data[225:450,1] = 1
    target_data[450:675,2] = 1

    
    # Create new sheet
    sheet = workbook.create_sheet(title='Правильна класифікація') 


    # Centring a text in cell
    for row in range(676):
        for column in range(28):
            sheet.cell(row=row+1, column=column+1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')


    # Add table of symbols name
    sheet.cell(row=1, column=1, value='Символ №1').fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    sheet.cell(row=1, column=2, value='Символ №2').fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    sheet.cell(row=1, column=3, value='Символ №3').fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")


    # Add color and value to matrix
    for row in range(675):
        sheet.cell(row=row+2, column=1, value=target_data[row,0]).fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        sheet.cell(row=row+2, column=2, value=target_data[row,1]).fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        sheet.cell(row=row+2, column=3, value=target_data[row,2]).fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")


    # Return noise matrix   
    workbook.save(path_to_xlsx) 

    return target_data