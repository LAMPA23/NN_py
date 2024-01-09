import numpy as np 
import openpyxl
import math


def apply_centroids(path_to_xlsx, workbook, sheet_name, centroids, noise_matrix, corect_ans):
    
    # Create void matrix
    Euclidean_distance = np.zeros((675,3))
    ans = np.zeros((675,3)).astype(int)
    err = np.zeros(3).astype(int)


    # Calculate Euclidean distance
    for row in range(675):
        for symbol in range(3): 
            for column in range(28):     
                Euclidean_distance[row, symbol] += (centroids[symbol][column] - noise_matrix[row,column])**2
            Euclidean_distance[row, symbol] = math.sqrt(Euclidean_distance[row, symbol])

        # Chack how Euclidean distance is less
        min_evclid_distance = min(Euclidean_distance[row, 0], Euclidean_distance[row, 1], Euclidean_distance[row, 2])
        for cnt in range(3):
            if Euclidean_distance[row, cnt] == min_evclid_distance:
                ans[row,cnt] = 1
                

    # Error check
    for row in range(225):
        for symbol in range(3):
            if not np.all(corect_ans[symbol * 225 + row, :] == ans[symbol * 225 + row, :]):
                err[symbol] += 1
    


    # Excel print ----------------------------

    # Create new sheet
    sheet = workbook.create_sheet(title=sheet_name) 

    # Centring a text and add value in cell)
    for row in range(675):
        for column in range(6):
            if column < 3:
                sheet.cell(row=row+1, column=column+1, value=ans[row,column]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            else:
                sheet.cell(row=row+1, column=column+1, value=corect_ans[row,column - 3]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            if column < 3:
                sheet.cell(row=row+1, column=column+1).fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            else:
                sheet.cell(row=row+1, column=column+1).fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

        if not np.all(corect_ans[row,:] == ans[row,:]):
            for cnt in range(6):
                sheet.cell(row=row+1, column=cnt+1).fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")           


    # Add table tilte and color to table tilte
    sheet.cell(row=1, column=1, value='Правильне значення №1').fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    sheet.cell(row=1, column=2, value='Правильне значення №2').fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    sheet.cell(row=1, column=3, value='Правильне значення №3').fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    sheet.cell(row=1, column=4, value='Отримане значення №1').fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    sheet.cell(row=1, column=5, value='Отримане значення №2').fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    sheet.cell(row=1, column=6, value='Отримане значення №3').fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    
    for cnt in range(3):
        sheet.cell(row=1, column=cnt+8, value=f'К-сть помилок №{cnt+1}').fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        sheet.cell(row=2, column=cnt+8, value=err[cnt]).fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        sheet.cell(row=1, column=cnt+8).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet.cell(row=2, column=cnt+8).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # Save data to Excel   
    workbook.save(path_to_xlsx)