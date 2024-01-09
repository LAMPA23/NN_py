import openpyxl
import numpy as np

def apply_KL(path_to_xlsx, workbook, sheet_name, mixed_noise_matrix, KM_centroids, corect_ans):
    
    # Create zeros matrix
    layers = np.zeros(3)
    ans = np.zeros((675, 3))
    err = np.zeros(3).astype(int)
    

    # Calculated result of network work 
    for row in range(675):
        for column in range(28):
            layers[0] += KM_centroids[0][column] * mixed_noise_matrix[row,column]
            layers[1] += KM_centroids[1][column] * mixed_noise_matrix[row,column]
            layers[2] += KM_centroids[2][column] * mixed_noise_matrix[row,column]
        ans[row, np.argmax(layers)] = 1


    # Error check
    for row in range(225):
        for symbol in range(3):
            if not np.all(corect_ans[symbol * 225 + row, :] == ans[symbol * 225 + row, :]):
                err[symbol] += 1

    # Create new sheet
    sheet = workbook.create_sheet(title=sheet_name)


    # Centring a text and add value in cell
    for row in range(675):
        for column in range(6):
            if column < 3:
                sheet.cell(row=row+1, column=column+1, value=ans[row,column]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            else:
                sheet.cell(row=row+1, column=column+1, value=corect_ans[row,column - 3]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            if column < 3:
                sheet.cell(row=row+1, column=column+1).fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            else:
                sheet.cell(row=row+1, column=column+1).fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

        if not np.all(corect_ans[row,:] == ans[row,:]):
            for cnt in range(6):
                sheet.cell(row=row+1, column=cnt+1).fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")           


    # Add table tilte and color to table tilte
    sheet.cell(row=1, column=1, value='Правильне значення №1').fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    sheet.cell(row=1, column=2, value='Правильне значення №2').fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    sheet.cell(row=1, column=3, value='Правильне значення №3').fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    sheet.cell(row=1, column=4, value='Отримане значення №1').fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    sheet.cell(row=1, column=5, value='Отримане значення №2').fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    sheet.cell(row=1, column=6, value='Отримане значення №3').fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    
    for cnt in range(3):
        sheet.cell(row=1, column=cnt+8, value=f'К-сть помилок №{cnt+1}').fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        sheet.cell(row=2, column=cnt+8, value=err[cnt]).fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        sheet.cell(row=1, column=cnt+8).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet.cell(row=2, column=cnt+8).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')


    # Save data to Excel   
    workbook.save(path_to_xlsx)

    

